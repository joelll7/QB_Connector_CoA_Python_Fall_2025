from __future__ import annotations

from pathlib import Path  # Filesystem path management
from typing import List  # Concrete list type for return value

from openpyxl import load_workbook  # Excel file loader

# Use absolute import so it works both as module and script from project root
from .models import Account  # Domain model used as output


def extract_account(workbook_path: Path) -> List[Account]:
    """Return payment terms parsed from the Excel workbook.

    Students should implement this function using ``openpyxl``. It must read the
    ``payment_terms`` worksheet, build :class:`~payment_terms_cli.models.PaymentTerm`
    instances with ``source="excel"``, and raise :class:`FileNotFoundError`
    if the workbook cannot be located.
    """

    workbook_path = Path(workbook_path)  # Ensure we have a Path instance
    if not workbook_path.exists():  # Validate the file exists
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    # Open in read-only mode for performance and safety; use cell values only
    workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook["chartofaccount"]  # Access the required worksheet by name
    except KeyError as exc:
        workbook.close()  # Close workbook before raising
        raise ValueError("Worksheet 'chartofaccount' not found in workbook") from exc

    rows = sheet.iter_rows(values_only=True)  # Iterate rows as tuples of raw values
    headers_row = next(rows, None)  # First row should contain column headers
    if headers_row is None:  # Empty sheet edge case
        workbook.close()
        return []

    # Build a mapping from header name to its column index
    headers = [
        str(header).strip() if header is not None else "" for header in headers_row
    ]
    header_index = {header: idx for idx, header in enumerate(headers)}

    def _value(row, column_name: str):  # Helper to safely access a column
        idx = header_index.get(column_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    terms: List[Account] = []  # Accumulator for valid terms
    try:
        for row in rows:  # Iterate over each data row
            raw_id = _value(row, "ID")  # Expected ID column (e.g., number of days)
            num = _value(row, "Number")  # Expected Number column

            name = _value(row, "Name")  # Expected Name column
            if name is None:
                continue  # Skip rows without a name
            name_str = str(name).strip()
            if not name_str:
                continue  # Skip blank names
            type = _value(row, "Type")  # Expected Type column
            if type is None:
                continue  # Skip rows without a type
            type_str = str(type).strip()
            if not type_str:
                continue  # Skip blank types
            if raw_id in (None, ""):
                continue  # Skip rows without an ID
            if num in (None, ""):
                continue  # Skip rows without a Number

            try:
                record_id = str(
                    raw_id
                ).strip()  # Normalise numerics (e.g., 30.0 -> "30")
                number = str(num).strip()
            except (TypeError, ValueError):
                record_id = str(raw_id).strip()  # Fallback to string trimming
                number = str(num).strip()

            if not record_id:
                continue  # Skip empty/invalid IDs
            if not number:
                continue  # Skip empty/invalid Numbers

            # Construct the domain object tagged as sourced from Excel
            terms.append(
                Account(
                    number=number,
                    id=record_id,
                    name=name_str,
                    AccountType=type_str,
                    source="excel",
                )
            )
    finally:
        workbook.close()  # Always close the workbook handle

    return terms  # Return the extracted list of payment terms


__all__ = ["extract_account"]  # Public API

if __name__ == "__main__":  # pragma: no cover - manual invocation
    import sys

    # Allow running as a script: poetry run python payment_terms_cli/excel_reader.py
    try:
        terms = extract_account(
            Path(
                "C:\\Users\\KieblesD\\Project\\QB_Connector_CoA_Python_Fall_2025\\company_data.xlsx"
            )
        )
        for term in terms:
            print(term)
    except Exception as e:
        print(f"Error: {e}")
        print("Usage: python src/excel_reader.py <path-to-workbook.xlsx>")
        sys.exit(1)
